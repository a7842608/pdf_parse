import os

import openpyxl

from utils.db import ConnectDatabase

# pat = r'C:\Users\86138\Desktop\20200905房小团\pdf_parse\utils\002pdf_excel.xlsx'

db = ConnectDatabase('ohop')

def readfile():
    rootdir = r'.\yhjg'
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    for i in range(0, len(list)):
        # for i in range(3, 4):
        pos = os.path.join(rootdir, list[i])
        # print(pos)
        if os.path.isfile(pos):
            # 你想对文件的操作
            pt = os.path.abspath(pos) # 绝对路径
            # print(pt)
            basename = os.path.basename(pos)
            name_id = basename.split('_')[0]
            print(name_id) # pid

            # 打开excel文件,获取工作簿对象
            wb = openpyxl.load_workbook(pt)
            # 从表单中获取单元格的内容
            ws = wb.active  # 当前活跃的表单
            try:
                for row in ws[1:int(ws.max_row) - 1]:  # 打印 2-5行中所有单元格中的值
                    try:
                        xuhao = str(row[0].value)
                    except:
                        xuhao = ''
                    try:
                        gfdjh = str(row[1].value)
                    except:
                        gfdjh = ''
                    try:
                        gfrxm = str(row[2].value)
                    except:
                        gfrxm = ''
                    try:
                        gfrzhengjian = str(row[3].value)
                    except:
                        gfrzhengjian = ''
                    try:
                        jia = str(row[4].value)
                    except:
                        jia = ''
                    try:
                        chadang = str(row[5].value)
                    except:
                        chadang = ''
                    try:
                        qita = str(row[6].value)
                    except:
                        qita = ''
                    try:
                        qitacheng = str(row[7].value)
                    except:
                        qitacheng = ''

                    sql = "insert into yixiangdengji(pid,shunxu,gfdengjihao,gfrxingming,goufangrenID," \
                          "jiatingleixing,bianhao,qitarenxingming,qitarenID) values('{}','{}','{}'," \
                          "'{}','{}','{}','{}','{}','{}');".format(name_id, xuhao, gfdjh, gfrxm, gfrzhengjian,
                                                                   jia, chadang, qita, qitacheng
                                                                   )
                    db.run(sql)
            except:
                print(pt)
                continue


if __name__ == '__main__':

    readfile()

    db.close()

# print(ws.max_column)
# print(ws.max_row)

# # 打开excel文件,获取工作簿对象
            # wb = openpyxl.load_workbook(pat)
            # # 从表单中获取单元格的内容
            # ws = wb.active  # 当前活跃的表单
            #
            # # print(ws.max_column)
            # print(ws.max_row)
            #
            # for row in ws[1:int(ws.max_row) - 1]:  # 打印 2-5行中所有单元格中的值
            #     print(row, type(row))
            #     # for cell in row:
            #     #     print(cell.value)
            #     # print(row[1].value)
            #     xuhao = str(row[0].value)
            #     gfdjh = str(row[1].value)
            #     gfrxm = str(row[2].value)
            #     gfrzhengjian = str(row[3].value)
            #     jia = str(row[4].value)
            #     chadang = str(row[5].value)
            #     qita = str(row[6].value)
            #     qitacheng = str(row[7].value)